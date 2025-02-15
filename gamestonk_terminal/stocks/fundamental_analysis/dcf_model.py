""" DCF Model """
__docformat__ = "numpy"

from urllib.request import urlopen
from typing import List, Union
from zipfile import ZipFile
from io import BytesIO

from sklearn.linear_model import LinearRegression
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
from openpyxl import worksheet
import yfinance as yf
import pandas as pd
import requests


opts = Union[int, str, float]


def string_float(string: str):
    """Numpy vectorize function to convert strings to floats"""
    return float(string.replace(",", ""))


def insert_row(name: str, index: str, df: pd.DataFrame, row_value: List[str]):
    """Allows a row to be inserted after a given row in a pandas dataframe"""
    pd.options.mode.chained_assignment = None
    if name not in df.index:
        row_number = df.index.get_loc(index) + 1
        df1 = df[0:row_number]
        df2 = df[row_number:]
        df1.loc[name] = row_value
        df_result = pd.concat([df1, df2])
        return df_result
    return df


def set_cell(
    ws: worksheet,
    cell: str,
    text: opts = None,
    font: str = None,
    border: str = None,
    fill: str = None,
    alignment: str = None,
    num_form: str = None,
):
    """Sets the value of the cell to given text and formats based on specified arguments"""
    if text:
        ws[cell] = text
    if font:
        ws[cell].font = font
    if border:
        ws[cell].border = border
    if fill:
        ws[cell].fill = fill
    if alignment:
        ws[cell].alignment = alignment
    if num_form:
        ws[cell].number_format = num_form


def get_rf():
    """Uses the fiscaldata.gov API to get most recent T-Bill rate"""
    base = "https://api.fiscaldata.treasury.gov/services/api/fiscal_service"
    end = "/v2/accounting/od/avg_interest_rates"
    filters = "?filter=security_desc:eq:Treasury Bills&sort=-record_date"
    response = requests.get(base + end + filters)
    latest = response.json()["data"][0]
    return latest["avg_interest_rate_amt"]


def get_fama_raw():
    """Gets base Fama French data to calculate risk"""
    with urlopen(
        "http://mba.tuck.dartmouth.edu/pages/faculty/ken.french/ftp/F-F_Research_Data_Factors_CSV.zip"
    ) as url:

        # Download Zipfile and create pandas DataFrame
        with ZipFile(BytesIO(url.read())) as zipfile:
            with zipfile.open("F-F_Research_Data_Factors.CSV") as zip_open:
                df = pd.read_csv(
                    zip_open,
                    header=0,
                    names=["Date", "MKT-RF", "SMB", "HML", "RF"],
                    skiprows=3,
                )

    df = df[df["Date"].apply(lambda x: len(str(x).strip()) == 6)]
    df["Date"] = df["Date"].astype(str) + "01"
    df["Date"] = pd.to_datetime(df["Date"], format="%Y%m%d")
    df["MKT-RF"] = pd.to_numeric(df["MKT-RF"], downcast="float")
    df["SMB"] = pd.to_numeric(df["SMB"], downcast="float")
    df["HML"] = pd.to_numeric(df["HML"], downcast="float")
    df["RF"] = pd.to_numeric(df["RF"], downcast="float")
    df["MKT-RF"] = df["MKT-RF"] / 100
    df["SMB"] = df["SMB"] / 100
    df["HML"] = df["HML"] / 100
    df["RF"] = df["RF"] / 100
    df = df.set_index("Date")
    return df


def get_historical_5(ticker: str):
    """Get 5 year monthly historical performance for a ticker with dividends filtered"""
    tick = yf.Ticker(ticker)
    df = tick.history(period="5y", interval="1mo")
    df = df[df.index.to_series().apply(lambda x: x.day == 1)]
    df = df.drop(["Dividends", "Stock Splits"], axis=1)
    df = df.dropna()
    return df


def get_fama_coe(ticker: str):
    """Use Fama and French to get the cost of equity for a company"""
    df_f = get_fama_raw()
    df_h = get_historical_5(ticker)
    df = df_h.join(df_f)
    df = df.dropna()
    df["Monthly Return"] = df["Close"].pct_change()
    df["Excess Monthly Return"] = df["Monthly Return"] - df["RF"]
    df = df.dropna()
    x = df[["MKT-RF", "SMB", "HML"]]
    y = df["Excess Monthly Return"]

    model = LinearRegression().fit(x, y)
    coefs = model.coef_
    return (
        df["RF"].mean()
        + coefs[0] * df["MKT-RF"].mean()
        + coefs[1] * df["SMB"].mean()
        + coefs[2] * df["HML"].mean()
    ) * 12


letters = [
    "A",
    "B",
    "C",
    "D",
    "E",
    "F",
    "G",
    "H",
    "I",
    "J",
    "K",
    "L",
    "M",
    "N",
    "O",
    "P",
    "Q",
    "R",
    "S",
    "T",
    "U",
    "V",
    "W",
    "X",
    "Y",
    "Z",
    "AA",
    "AB",
    "AC",
    "AD",
    "AE",
    "AF",
    "AG",
    "AH",
    "AI",
    "AJ",
    "AK",
    "AL",
    "AM",
    "AN",
    "AO",
    "AP",
    "AQ",
    "AR",
    "AS",
    "AT",
    "AU",
    "AV",
    "AW",
    "AX",
    "AY",
    "AZ",
]
non_gaap_is = [
    "Revenue Growth",
    "Net Income Common",
    "Net Income Growth",
    "Shares Outstanding (Basic)",
    "Shares Outstanding (Diluted)",
    "Shares Change",
    "EPS (Basic)",
    "EPS (Diluted)",
    "EPS Growth",
    "Free Cash Flow Per Share",
    "Dividend Per Share",
    "Dividend Growth",
    "Gross Margin",
    "Operating Margin",
    "Profit Margin",
    "Free Cash Flow Margin",
    "Effective Tax Rate",
    "EBITDA",
    "EBITDA Margin",
    "EBIT",
    "EBIT Margin",
    "Operating Expenses",
    "Pretax Income",
]
gaap_is = [
    "Revenue",
    "Cost of Revenue",
    "Gross Profit",
    "Selling, General & Admin",
    "Research & Development",
    "Other Operating Expenses",
    "Operating Income",
    "Interest Expense / Income",
    "Other Expense / Income",
    "Income Tax",
    "Net Income",
    "Preferred Dividends",
]
non_gaap_bs = [
    "Cash Growth",
    "Debt Growth",
    "Net Cash / Debt",
    "Net Cash / Debt Growth",
    "Net Cash Per Share",
    "Working Capital",
    "Book Value Per Share",
    "Total Debt",
]
gaap_bs = [
    "Cash & Equivalents",
    "Short-Term Investments",
    "Cash & Cash Equivalents",
    "Receivables",
    "Inventory",
    "Other Current Assets",
    "Total Current Assets",
    "Property, Plant & Equipment",
    "Long-Term Investments",
    "Goodwill and Intangibles",
    "Other Long-Term Assets",
    "Total Long-Term Assets",
    "Total Assets",
    "Accounts Payable",
    "Deferred Revenue",
    "Current Debt",
    "Other Current Liabilities",
    "Total Current Liabilities",
    "Long-Term Debt",
    "Other Long-Term Liabilities",
    "Total Long-Term Liabilities",
    "Total Liabilities",
    "Common Stock",
    "Retained Earnings",
    "Comprehensive Income",
    "Shareholders' Equity",
    "Total Liabilities and Equity",
]
non_gaap_cf = [
    "Operating Cash Flow Growth",
    "Free Cash Flow Growth",
    "Free Cash Flow Margin",
    "Free Cash Flow Per Share",
    "Free Cash Flow",
]

gaap_cf = [
    "Net Income",
    "Depreciation & Amortization",
    "Share-Based Compensation",
    "Other Operating Activities",
    "Operating Cash Flow",
    "Capital Expenditures",
    "Acquisitions",
    "Change in Investments",
    "Other Investing Activities",
    "Investing Cash Flow",
    "Dividends Paid",
    "Share Issuance / Repurchase",
    "Debt Issued / Paid",
    "Other Financing Activities",
    "Financing Cash Flow",
    "Net Cash Flow",
]

sum_rows = [
    "Gross Profit",
    "Operating Income",
    "Net Income",
    "Cash & Cash Equivalents",
    "Total Current Assets",
    "Total Long-Term Assets",
    "Total Assets",
    "Total Current Liabilities",
    "Total Long-Term Liabilities",
    "Total Liabilities",
    "Shareholders' Equity",
    "Total Liabilities and Equity",
    "Operating Cash Flow",
    "Investing Cash Flow",
    "Financing Cash Flow",
    "Net Cash Flow",
]

bold_font = Font(bold=True)
thin_border_top = Border(top=Side(style="thin"))
thin_border_bottom = Border(bottom=Side(style="thin"))

thin_border_nl = Border(
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

thin_border_nr = Border(
    left=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

green_bg = PatternFill(fgColor="7fe5cd", fill_type="solid")

center = Alignment(horizontal="center")

red = Font(color="FF0000")

fmt_acct = "_($* #,##0.00_);[Red]_($* (#,##0.00);_($* -_0_0_);_(@"

headers = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/71.0.3578.98 Safari/537.36"
    )
}

tickers = [
    "AEIS",
    "AEL",
    "AEM",
    "AEMD",
    "AENZ",
    "AEO",
    "AEP",
    "AER",
    "AERI",
    "AES",
]
